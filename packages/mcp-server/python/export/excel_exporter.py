"""Markdown -> Excel (.xlsx) exporter with IPA 4-sheet structure."""

import json
import re
import sys
from pathlib import Path

import mistune
import yaml
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from openpyxl.styles import Font as OpenpyxlFont, PatternFill

from .shared_styles import (
    HEADER_FONT, DATA_FONT, TITLE_FONT, SUBTITLE_FONT,
    HEADER_BG, ALT_ROW_BG, THIN_BORDER,
    HEADER_ALIGN, DATA_ALIGN, CENTER_ALIGN,
    jp_column_width,
)

# Revision marker prefixes and their styles (朱書き mode)
REVISION_MARKERS = ["【新規】", "【変更】", "【削除】"]
REVISION_FILLS = {
    "【新規】": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "【変更】": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "【削除】": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}
REVISION_FONTS = {
    "【新規】": OpenpyxlFont(name="MS Gothic", size=10, color="CC0000"),
    "【変更】": OpenpyxlFont(name="MS Gothic", size=10, bold=True),
    "【削除】": OpenpyxlFont(name="MS Gothic", size=10, color="808080", strikethrough=True),
}

# Frontmatter regex
FRONTMATTER_RE = re.compile(r"^---\n([\s\S]*?)\n---\n([\s\S]*)$")
# Table row regex
TABLE_ROW_RE = re.compile(r"^\|(.+)\|$")
TABLE_SEP_RE = re.compile(r"^\|[\s\-:|]+\|$")


def parse_markdown(content: str) -> dict:
    """Parse MD content into frontmatter, headings, and tables."""
    meta = {}
    body = content

    fm_match = FRONTMATTER_RE.match(content)
    if fm_match:
        meta = yaml.safe_load(fm_match.group(1)) or {}
        body = fm_match.group(2)

    headings = []
    tables = []
    current_table = None

    for line in body.split("\n"):
        # Headings
        h_match = re.match(r"^(#{1,4})\s+(.+)$", line)
        if h_match:
            headings.append({"level": len(h_match.group(1)), "text": h_match.group(2).strip()})
            if current_table:
                tables.append(current_table)
                current_table = None
            continue

        # Skip separator rows
        if TABLE_SEP_RE.match(line.strip()):
            continue

        # Table rows
        row_match = TABLE_ROW_RE.match(line.strip())
        if row_match:
            cells = [c.strip() for c in row_match.group(1).split("|")]
            if current_table is None:
                current_table = {"header": cells, "rows": []}
            else:
                current_table["rows"].append(cells)
        else:
            if current_table:
                tables.append(current_table)
                current_table = None

    if current_table:
        tables.append(current_table)

    return {"meta": meta, "headings": headings, "tables": tables, "body": body}


def create_cover_sheet(wb: Workbook, meta: dict) -> None:
    """Sheet 1: 表紙 (cover page)."""
    ws = wb.active
    ws.title = "表紙"
    ws.sheet_properties.tabColor = "203864"

    ws.merge_cells("B4:F4")
    cell = ws["B4"]
    cell.value = meta.get("doc_type", "設計書")
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN

    info = [
        ("プロジェクト名", meta.get("project_name", "")),
        ("ドキュメント種別", meta.get("doc_type", "")),
        ("版数", meta.get("version", "1.0")),
        ("言語", meta.get("language", "ja")),
    ]
    for i, (label, value) in enumerate(info, start=7):
        ws[f"C{i}"].value = label
        ws[f"C{i}"].font = SUBTITLE_FONT
        ws[f"D{i}"].value = value
        ws[f"D{i}"].font = DATA_FONT

    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40


def create_history_sheet(wb: Workbook) -> None:
    """Sheet 2: 更新履歴."""
    ws = wb.create_sheet("更新履歴")
    ws.sheet_properties.tabColor = "4472C4"
    headers = ["版数", "更新日", "更新者", "更新内容"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.border = THIN_BORDER
        cell.alignment = HEADER_ALIGN
    ws.cell(row=2, column=1, value="1.0")
    ws.cell(row=2, column=4, value="初版作成")


def create_toc_sheet(wb: Workbook, headings: list) -> None:
    """Sheet 3: 目次."""
    ws = wb.create_sheet("目次")
    ws.sheet_properties.tabColor = "4472C4"
    ws.cell(row=1, column=1, value="目次").font = TITLE_FONT

    for i, h in enumerate(headings, start=3):
        indent = "  " * (h["level"] - 1)
        ws.cell(row=i, column=1, value=f"{indent}{h['text']}")
        ws.cell(row=i, column=1).font = DATA_FONT


def create_content_sheet(wb: Workbook, table: dict, sheet_name: str) -> None:
    """Create a content sheet from a parsed table."""
    ws = wb.create_sheet(sheet_name[:31])  # Excel 31-char limit

    # Header row
    for col, h in enumerate(table["header"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.border = THIN_BORDER
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col)].width = jp_column_width(h)

    # Data rows
    for row_idx, row in enumerate(table["rows"], 2):
        # Detect revision marker from first cell
        row_marker = ""
        for marker in REVISION_MARKERS:
            if row and row[0].strip().startswith(marker):
                row_marker = marker
                break

        for col_idx, val in enumerate(row, 1):
            # Strip marker prefix from display value
            display_val = val
            if row_marker and col_idx == 1:
                display_val = val.strip().removeprefix(row_marker).strip()

            cell = ws.cell(row=row_idx, column=col_idx, value=display_val)
            cell.border = THIN_BORDER
            cell.alignment = DATA_ALIGN

            if row_marker:
                cell.fill = REVISION_FILLS.get(row_marker, ALT_ROW_BG)
                cell.font = REVISION_FONTS.get(row_marker, DATA_FONT)
            else:
                cell.font = DATA_FONT
                if row_idx % 2 == 0:
                    cell.fill = ALT_ROW_BG

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-adjust widths from data
    for col_idx in range(1, len(table["header"]) + 1):
        max_width = jp_column_width(table["header"][col_idx - 1])
        for row in table["rows"]:
            if col_idx - 1 < len(row):
                max_width = max(max_width, jp_column_width(row[col_idx - 1]))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 50)

    # Add data validation for known enum columns
    header_lower = [h.strip() for h in table["header"]]
    for col_idx, h in enumerate(header_lower, 1):
        if "処理分類" in h:
            dv = DataValidation(type="list", formula1='"入力,照会,帳票,バッチ"')
            dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}200")
            ws.add_data_validation(dv)
        elif "優先度" in h or "難易度" in h:
            dv = DataValidation(type="list", formula1='"高,中,低"')
            dv.add(f"{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}200")
            ws.add_data_validation(dv)

    # Print setup
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"


def export(content: str, doc_type: str, output_path: str, project_name: str = "") -> dict:
    """Main export entry point: MD -> Excel."""
    parsed = parse_markdown(content)
    meta = {**parsed["meta"], "doc_type": doc_type, "project_name": project_name}

    wb = Workbook()
    create_cover_sheet(wb, meta)
    create_history_sheet(wb)
    create_toc_sheet(wb, parsed["headings"])

    # Create content sheets from tables
    for i, table in enumerate(parsed["tables"]):
        name = f"本文{i + 1}" if i > 0 else "本文"
        create_content_sheet(wb, table, name)

    # If no tables found, create empty content sheet
    if not parsed["tables"]:
        ws = wb.create_sheet("本文")
        ws.cell(row=1, column=1, value="コンテンツなし").font = DATA_FONT

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    file_size = Path(output_path).stat().st_size
    return {"success": True, "file_path": output_path, "file_size": file_size}


if __name__ == "__main__":
    input_data = json.loads(sys.stdin.read())
    result = export(
        input_data["content"],
        input_data["doc_type"],
        input_data["output_path"],
        input_data.get("project_name", ""),
    )
    print(json.dumps(result))
