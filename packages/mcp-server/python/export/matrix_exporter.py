"""CRUD matrix and traceability matrix -> Excel (.xlsx) exporter."""

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .shared_styles import (
    HEADER_FONT, DATA_FONT, TITLE_FONT, SUBTITLE_FONT,
    HEADER_BG, ALT_ROW_BG, ACCENT_BG, THIN_BORDER,
    HEADER_ALIGN, DATA_ALIGN, CENTER_ALIGN,
    jp_column_width,
)

# Regex for markdown table rows
TABLE_ROW_RE = re.compile(r"^\|(.+)\|$")
TABLE_SEP_RE = re.compile(r"^\|[\s\-:|]+\|$")

# CRUD cell highlight colors
CRUD_COLORS = {
    "C": "C6EFCE",  # green
    "R": "BDD7EE",  # blue
    "U": "FFE699",  # yellow
    "D": "F4CCCC",  # red
}


def _parse_table(content: str) -> dict | None:
    """Extract first markdown table from content as {header, rows}."""
    header = None
    rows = []
    for line in content.split("\n"):
        line = line.strip()
        if TABLE_SEP_RE.match(line):
            continue
        m = TABLE_ROW_RE.match(line)
        if m:
            cells = [c.strip() for c in m.group(1).split("|")]
            if header is None:
                header = cells
            else:
                rows.append(cells)
        elif header is not None:
            break  # end of table
    if header:
        return {"header": header, "rows": rows}
    return None


def _write_matrix_sheet(wb: Workbook, table: dict, sheet_name: str, is_crud: bool) -> None:
    """Write a matrix table to an Excel sheet with formatting."""
    ws = wb.create_sheet(sheet_name[:31])
    from openpyxl.styles import PatternFill

    # Header row
    for col, h in enumerate(table["header"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_BG
        cell.border = THIN_BORDER
        cell.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(col)].width = jp_column_width(h)

    # Data rows
    for row_idx, row in enumerate(table["rows"], 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col_idx > 2 else DATA_ALIGN

            # Color CRUD cells
            if is_crud and col_idx > 2 and val.strip():
                for letter in "CRUD":
                    if letter in val.upper():
                        cell.fill = PatternFill(
                            start_color=CRUD_COLORS[letter],
                            end_color=CRUD_COLORS[letter],
                            fill_type="solid",
                        )
                        break
            # Color ○ cells in traceability
            elif not is_crud and col_idx > 2 and "○" in val:
                cell.fill = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )

            if row_idx % 2 == 0 and (not cell.fill or cell.fill.fill_type is None):
                cell.fill = ALT_ROW_BG

    ws.freeze_panes = "C2"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "landscape"


def export_matrix(
    content: str,
    matrix_type: str,
    output_path: str,
    project_name: str = "",
) -> dict:
    """Export CRUD or traceability matrix markdown to Excel.

    Args:
        content: Markdown content containing the matrix table
        matrix_type: "crud-matrix" or "traceability-matrix"
        output_path: Where to save the .xlsx file
        project_name: Project name for cover sheet
    """
    table = _parse_table(content)
    if not table:
        return {"error": "No markdown table found in content"}

    is_crud = matrix_type == "crud-matrix"
    wb = Workbook()

    # Cover sheet
    ws = wb.active
    ws.title = "表紙"
    ws.sheet_properties.tabColor = "203864"
    ws.merge_cells("B4:F4")
    cell = ws["B4"]
    cell.value = "CRUD図" if is_crud else "トレーサビリティマトリックス"
    cell.font = TITLE_FONT
    cell.alignment = CENTER_ALIGN

    info = [
        ("プロジェクト名", project_name),
        ("ドキュメント種別", cell.value),
        ("版数", "1.0"),
    ]
    for i, (label, value) in enumerate(info, start=7):
        ws[f"C{i}"].value = label
        ws[f"C{i}"].font = SUBTITLE_FONT
        ws[f"D{i}"].value = value
        ws[f"D{i}"].font = DATA_FONT
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40

    # Matrix sheet
    sheet_name = "CRUD図" if is_crud else "トレーサビリティ"
    _write_matrix_sheet(wb, table, sheet_name, is_crud)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    file_size = Path(output_path).stat().st_size
    return {"success": True, "file_path": output_path, "file_size": file_size}
