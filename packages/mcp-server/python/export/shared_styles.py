"""Shared JP styling constants for Excel and PDF exports."""

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# -- Fonts --
HEADER_FONT = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
DATA_FONT = Font(name="MS Gothic", size=10)
TITLE_FONT = Font(name="Meiryo", size=16, bold=True)
SUBTITLE_FONT = Font(name="Meiryo", size=12, bold=True)

# -- Colors --
HEADER_BG = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
ALT_ROW_BG = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
ACCENT_BG = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# -- Borders --
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# -- Alignment --
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# -- PDF CSS --
PDF_CSS = """
@page {
    size: A4 landscape;
    margin: 15mm 10mm;
    @bottom-center { content: "ページ " counter(page) " / " counter(pages); font-size: 9pt; }
}
body { font-family: "Noto Sans JP", "Meiryo", sans-serif; font-size: 10pt; line-height: 1.6; }
h1 { font-size: 18pt; color: #203864; border-bottom: 2px solid #203864; padding-bottom: 4pt; }
h2 { font-size: 14pt; color: #203864; margin-top: 12pt; }
h3 { font-size: 12pt; color: #4472C4; }
table { border-collapse: collapse; width: 100%; margin: 8pt 0; page-break-inside: auto; }
th { background-color: #203864; color: white; font-weight: bold; padding: 6pt 8pt;
     border: 1px solid #203864; text-align: center; }
td { padding: 4pt 8pt; border: 1px solid #D3D3D3; vertical-align: top; }
tr:nth-child(even) td { background-color: #F2F2F2; }
.toc { page-break-after: always; }
.toc a { text-decoration: none; color: #203864; }
.cover { text-align: center; padding-top: 200px; page-break-after: always; }
.cover h1 { font-size: 28pt; border: none; }
"""


def jp_column_width(text: str) -> float:
    """Estimate column width for JP text (2-byte chars are ~1.8x wider)."""
    width = 0.0
    for ch in text:
        width += 2.2 if ord(ch) > 127 else 1.0
    return max(width + 2, 8.0)
