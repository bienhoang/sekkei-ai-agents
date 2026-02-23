"""Import Excel documents into Sekkei markdown format."""

import re
from typing import Optional

# openpyxl is already in requirements.txt
import openpyxl


# Column patterns for auto-detecting document type
DOC_TYPE_PATTERNS = {
    "functions-list": ["機能ID", "機能名", "大分類", "中分類"],
    "requirements": ["要件ID", "要件名", "REQ-"],
    "test-spec": ["テストケースID", "テスト対象", "UT-", "IT-", "ST-"],
    "test-evidence": ["エビデンスID", "EV-"],
    "basic-design": ["画面ID", "SCR-", "テーブルID", "TBL-"],
}


def detect_doc_type(headers: list[str]) -> Optional[str]:
    """Auto-detect document type from column headers."""
    header_text = " ".join(headers)
    best_match = None
    best_score = 0

    for doc_type, patterns in DOC_TYPE_PATTERNS.items():
        score = sum(1 for p in patterns if p in header_text)
        if score > best_score:
            best_score = score
            best_match = doc_type

    return best_match if best_score > 0 else None


def cell_value(cell) -> str:
    """Extract string value from cell, handling None and merged cells."""
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def import_excel(
    file_path: str,
    doc_type_hint: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> dict:
    """Import Excel file and convert to Sekkei markdown.

    Args:
        file_path: Path to .xlsx file
        doc_type_hint: Optional document type hint
        sheet_name: Optional specific sheet to import

    Returns:
        dict with: content, detected_doc_type, sheet_count, row_count, warnings
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    warnings = []
    all_content = []
    total_rows = 0

    sheets = [wb[sheet_name]] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets
    detected_doc_type = doc_type_hint

    for ws in sheets:
        rows = list(ws.iter_rows())
        if not rows:
            continue

        # Find header row (first row with >3 non-empty cells)
        header_row_idx = 0
        for i, row in enumerate(rows):
            non_empty = sum(1 for c in row if c.value is not None)
            if non_empty >= 3:
                header_row_idx = i
                break

        headers = [cell_value(c) for c in rows[header_row_idx]]

        # Auto-detect doc type from first sheet headers
        if not detected_doc_type:
            detected_doc_type = detect_doc_type(headers)

        # Build markdown table
        section_title = ws.title or "Sheet"
        md_lines = [f"## {section_title}", ""]

        # Header row
        md_lines.append("| " + " | ".join(h or "Column" for h in headers) + " |")
        md_lines.append("| " + " | ".join("---" for _ in headers) + " |")

        # Data rows
        data_rows = rows[header_row_idx + 1:]
        for row in data_rows:
            values = [cell_value(c) for c in row[: len(headers)]]
            if any(v for v in values):  # skip completely empty rows
                md_lines.append("| " + " | ".join(values) + " |")
                total_rows += 1

        all_content.append("\n".join(md_lines))

    if not detected_doc_type:
        warnings.append("Could not auto-detect doc type from column patterns")

    # Build frontmatter
    frontmatter = [
        "---",
        f'doc_type: {detected_doc_type or "unknown"}',
        'version: "1.0"',
        "language: ja",
        "status: draft",
        "---",
    ]

    content = "\n".join(frontmatter) + "\n\n" + "\n\n".join(all_content)

    wb.close()

    return {
        "content": content,
        "detected_doc_type": detected_doc_type,
        "sheet_count": len(sheets),
        "row_count": total_rows,
        "warnings": warnings,
    }
